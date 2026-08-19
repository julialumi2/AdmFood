"""
Importa o comparativo de preços do cardápio (iFood, 99Food, BeeFood,
Cardápio Web) de uma planilha .xlsx pra tabela preco_cardapio — usada pela
tela Cardápio, só de referência/visualização (não tem edição pelo sistema).

A planilha tem uma aba por grupo de loja ("Comparativo de Preços Art",
"... Tradiças", "... Açaí"), com uma linha por produto:
  coluna A = nome do produto (ou o nome de uma categoria, quando as colunas
             de preço da linha estão todas vazias — ex: "BEBIDAS", "PORÇÕES")
  colunas seguintes = preço em cada canal, pela ordem do cabeçalho da aba
    (nem toda aba tem BeeFood — só a de Artesanos)

Sempre apaga e reimporta a tabela inteira (substituir_precos_cardapio) —
rode de novo sempre que a planilha for atualizada.

Uso:
  python importar_precos_cardapio.py "caminho/da/planilha.xlsx"
"""

import sys

import openpyxl

from backend.armazenamento import inicializar_banco, substituir_precos_cardapio

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# Nome da aba na planilha -> nome de exibição na tela Cardápio.
LOJA_POR_ABA = {
    "Comparativo de Preços Art": "Hamburgueria Artesanos",
    "Comparativo de Preços Tradiças": "Tradiças",
    "Comparativo de Preços Açaí": "Açaí Na Lata",
}

# Nome da coluna do cabeçalho (case-insensitive, sem o "(R$)") -> coluna no banco.
CANAL_POR_CABECALHO = {
    "ifood": "ifood",
    "99 food": "food99",
    "beefood": "beefood",
    "cardápio web": "cardapio_web",
}


def _normalizar_cabecalho(texto):
    return (texto or "").replace("(R$)", "").strip().lower()


def _extrair_linhas_da_aba(aba, loja):
    cabecalho = [_normalizar_cabecalho(c.value) for c in next(aba.iter_rows(min_row=1, max_row=1))]
    colunas_canal = {}
    for indice, nome in enumerate(cabecalho):
        if nome in CANAL_POR_CABECALHO:
            colunas_canal[CANAL_POR_CABECALHO[nome]] = indice

    linhas = []
    categoria_atual = "Geral"
    ordem = 0
    for linha in aba.iter_rows(min_row=2):
        produto = (linha[0].value or "").strip() if linha[0].value else ""
        if not produto:
            continue

        precos = {}
        for coluna_banco, indice in colunas_canal.items():
            valor = linha[indice].value if indice < len(linha) else None
            precos[coluna_banco] = float(valor) if isinstance(valor, (int, float)) else None

        # Linha de categoria: só tem o nome na coluna A, nenhum preço em canal nenhum.
        if all(v is None for v in precos.values()):
            categoria_atual = produto
            continue

        ordem += 1
        linhas.append({
            "loja": loja,
            "categoria": categoria_atual,
            "produto": produto,
            "ifood": precos.get("ifood"),
            "food99": precos.get("food99"),
            "beefood": precos.get("beefood"),
            "cardapio_web": precos.get("cardapio_web"),
            "ordem": ordem,
        })
    return linhas


def importar(caminho_planilha):
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)

    todas_linhas = []
    for nome_aba, loja in LOJA_POR_ABA.items():
        aba = next((wb[n] for n in wb.sheetnames if n.strip() == nome_aba.strip()), None)
        if aba is None:
            print(f"⚠️  Aba \"{nome_aba}\" não encontrada na planilha — pulando {loja}.")
            continue
        linhas = _extrair_linhas_da_aba(aba, loja)
        todas_linhas.extend(linhas)
        print(f"✅ {loja}: {len(linhas)} produtos")

    substituir_precos_cardapio(todas_linhas)
    print(f"\nTotal importado: {len(todas_linhas)} produtos.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_precos_cardapio.py \"caminho/da/planilha.xlsx\"")
        sys.exit(1)

    inicializar_banco()
    importar(sys.argv[1])
