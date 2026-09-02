"""
Lê o comparativo de preços do cardápio (iFood, 99Food, BeeFood, Cardápio
Web) de uma planilha .xlsx — usado tanto pelo script de linha de comando
(importar_precos_cardapio.py) quanto pelo upload direto na tela Cardápio
(admin only, ver app.py).

A planilha tem uma aba por grupo de loja ("Comparativo de Preços Art",
"... Tradiças", "... Açaí"), com uma linha por produto:
  coluna A = nome do produto (ou o nome de uma categoria, quando as colunas
             de preço da linha estão todas vazias — ex: "BEBIDAS", "PORÇÕES")
  colunas seguintes = preço em cada canal, pela ordem do cabeçalho da aba
    (nem toda aba tem BeeFood — só a de Artesanos)
"""

import openpyxl

# Nome da aba na planilha -> nome de exibição na tela Cardápio (uma aba
# pode alimentar mais de uma loja — Tradiça ZN e Tradiça Simus vendem pelo
# mesmo preço hoje, então continuam saindo da mesma aba única da
# planilha, só que agora como duas lojas de verdade em vez de um
# "Tradiças" compartilhado, pra ficar igual o resto do sistema; concluído
# em 2026-09-02, a pedido da Julia).
LOJA_POR_ABA = {
    "Comparativo de Preços Art": "Hamburgueria Artesanos",
    "Comparativo de Preços Tradiças": ["Tradiça ZN", "Tradiça Simus"],
    "Comparativo de Preços Açaí": "Açaí Na Lata",
}

# Nome da coluna do cabeçalho (case-insensitive, sem o "(R$)") -> coluna no banco.
CANAL_POR_CABECALHO = {
    "ifood": "ifood",
    "99 food": "food99",
    "beefood": "beefood",
    "cardápio web": "cardapio_web",
}


def _normalizar_cabecalho(texto):
    return (texto or "").replace("(R$)", "").strip().lower()


def _extrair_linhas_da_aba(aba, loja):
    cabecalho = [_normalizar_cabecalho(c.value) for c in next(aba.iter_rows(min_row=1, max_row=1))]
    colunas_canal = {}
    for indice, nome in enumerate(cabecalho):
        if nome in CANAL_POR_CABECALHO:
            colunas_canal[CANAL_POR_CABECALHO[nome]] = indice

    linhas = []
    categoria_atual = "Geral"
    ordem = 0
    for linha in aba.iter_rows(min_row=2):
        produto = (linha[0].value or "").strip() if linha[0].value else ""
        if not produto:
            continue

        precos = {}
        for coluna_banco, indice in colunas_canal.items():
            valor = linha[indice].value if indice < len(linha) else None
            precos[coluna_banco] = float(valor) if isinstance(valor, (int, float)) else None

        # Linha de categoria: só tem o nome na coluna A, nenhum preço em canal nenhum.
        if all(v is None for v in precos.values()):
            categoria_atual = produto
            continue

        ordem += 1
        linhas.append({
            "loja": loja,
            "categoria": categoria_atual,
            "produto": produto,
            "ifood": precos.get("ifood"),
            "food99": precos.get("food99"),
            "beefood": precos.get("beefood"),
            "cardapio_web": precos.get("cardapio_web"),
            "ordem": ordem,
        })
    return linhas


def ler_precos_da_planilha(caminho_ou_arquivo):
    """caminho_ou_arquivo: caminho de arquivo (str) ou um objeto tipo-arquivo
    (ex: o stream de um upload) — openpyxl aceita os dois. Levanta
    ValueError se nenhuma das abas esperadas for encontrada."""
    wb = openpyxl.load_workbook(caminho_ou_arquivo, data_only=True)

    todas_linhas = []
    abas_encontradas = []
    for nome_aba, lojas in LOJA_POR_ABA.items():
        aba = next((wb[n] for n in wb.sheetnames if n.strip() == nome_aba.strip()), None)
        if aba is None:
            continue
        for loja in (lojas if isinstance(lojas, list) else [lojas]):
            abas_encontradas.append(loja)
            todas_linhas.extend(_extrair_linhas_da_aba(aba, loja))

    if not abas_encontradas:
        raise ValueError(
            "Nenhuma das abas esperadas foi encontrada na planilha "
            f"(procurado: {', '.join(LOJA_POR_ABA.keys())})."
        )

    return todas_linhas
